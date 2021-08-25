#!/usr/bin/env python
# -*- coding:utf-8 -*-
#for research progress
import urllib
import os
import time
import re
import logging
import requests
import urllib3
# json解析库,对应到lxml
import json
# 全局取消证书验证
import io
import sys
# 改变标准输出的默认编码
# utf-8中文乱码 有些表情print不进去
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')


import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import openpyxl
def douyin():
    video_count = 0  # 记录视频个数

    #for video_url in file_object:
    times=0
    promotion_id=3001194623569399560
    while(times<10000000):
        times=times+1
        t="https://api3-normal-c-lf.amemv.com/aweme/v2/shop/item/feed/?page=1&size=20&item_id=6850954270171303181&promotion_id=3423704625537248340&product_id=3423704625537248340&os_api=22&device_type=VOG-AL10&ssmix=a&manifest_version_code=110901&dpi=320&uuid=863064175923921&app_name=aweme&version_name=11.9.0&ts=1595558573&cpu_support64=false&storage_type=0&app_type=normal&ac=wifi&host_abi=armeabi-v7a&update_version_code=11909900&channel=tengxun_new&_rticket=1595558573617&device_platform=android&iid=3553214257185918&version_code=110900&mac_address=F8%3A63%3A3F%3A95%3AEF%3AF7&cdid=f8ae9f36-b51e-4915-8f2d-fc59650b6194&openudid=9283b6041f480e38&device_id=1855630942669101&resolution=900*1600&os_version=5.1.1&language=zh&device_brand=huawei&aid=1128&mcc_mnc=46007"
        video_url_first="https://api3-normal-c-lf.amemv.com/aweme/v2/shop/item/feed/?page=1&size=20&item_id=6852358686187326720&promotion_id="
        video_url_second="&product_id="
        video_url_third="&os_api=22&device_type=VOG-AL10&ssmix=a&manifest_version_code=110901&dpi=320&uuid=863064175923921&app_name=aweme&version_name=11.9.0&ts=1595558573&cpu_support64=false&storage_type=0&app_type=normal&ac=wifi&host_abi=armeabi-v7a&update_version_code=11909900&channel=tengxun_new&_rticket=1595558340439&device_platform=android&iid=3553214257185918&version_code=110900&mac_address=F8%3A63%3A3F%3A95%3AEF%3AF7&cdid=f8ae9f36-b51e-4915-8f2d-fc59650b6194&openudid=9283b6041f480e38&device_id=1855630942669101&resolution=900*1600&os_version=5.1.1&language=zh&device_brand=huawei&aid=1128&mcc_mnc=46007"

        promotion_id=promotion_id+1
        product_id=promotion_id
        product_id=product_id+1
        video_url=video_url_first+str(promotion_id)+video_url_second+str(product_id)+video_url_third

        if video_url.__sizeof__() > 10:
            data = openpyxl.load_workbook('douyin1.xlsx')
            # print(data.get_named_ranges())  # 输出工作页索引范围
            # print(data.get_sheet_names())  # 输出所有工作页的名称
            # 取第一张表
            sheetnames = data.get_sheet_names()
            table = data.get_sheet_by_name(sheetnames[0])
            table = data.active
            # print(table.title)  # 输出表名
            nrows = table.max_row  # 获得行数
            # ncolumns = table.max_column  # 获得行数
            index = 1

            print(index)
            req = urllib.request.urlopen(video_url).read()
            unicodestr = json.loads(req.decode())

            user_list = unicodestr.get("aweme_list")
            if user_list != None:
                for eve in user_list:
                    i=eve.get("anchor_info")
                    if(i!=None):
                        mechant_info = i.get("extra")
                        if (mechant_info != None):
                            table.cell(nrows + index, 1).value = str(mechant_info)
                            share_url= eve.get("share_url")
                            if(share_url!=None):
                                table.cell(nrows + index, 2).value = str(share_url)

                            descendants = eve.get("descendants")
                            if(descendants!=None):
                                platforms=descendants.get("platforms")
                                notify_msg = descendants.get("notify_msg")
                                if(platforms!=None):
                                   table.cell(nrows + index, 3).value = str(platforms)
                                if (notify_msg != None):
                                   table.cell(nrows + index, 4).value = str(platforms)

                            create_time = eve.get("create_time")#时间戳
                            if (create_time != None):
                                table.cell(nrows + index, 33).value = str(create_time)


                            author=eve.get("author")
                            if(author!=None):
                                apple_account=author.get("apple_account")
                                if(apple_account!=None):
                                    table.cell(nrows + index, 5).value = str(apple_account)
                                short_id=author.get("short_id")
                                if(short_id!=None):
                                    table.cell(nrows + index, 6).value = str(short_id)
                                temp=author.get("avatar_medium")
                                if(temp!=None):
                                    author_photos=temp.get("url_list")
                                    author_photo=author_photos[0]
                                    if(author_photo!=None):
                                        table.cell(nrows + index, 7).value = str(author_photo)
                                has_facebook_token=author.get("has_facebook_token")
                                if(has_facebook_token!=None):
                                    table.cell(nrows + index, 8).value = str(has_facebook_token)
                                is_ad_fake=author.get("is_ad_fake")
                                if(is_ad_fake!=None):
                                    table.cell(nrows + index, 9).value = str(is_ad_fake)
                                has_orders=author.get("has_orders")
                                if(has_orders!=None):
                                    table.cell(nrows + index, 10).value = str(has_orders)
                                live_agreement=author.get("live_agreement")
                                if(live_agreement!=None):
                                    table.cell(nrows + index, 11).value = str(live_agreement)
                                twitter_id=author.get("twitter_id")
                                if(twitter_id!=None):
                                    table.cell(nrows + index, 12).value = str(twitter_id)
                                duet_setting=author.get("duet_setting")
                                if(duet_setting!=None):
                                    table.cell(nrows + index, 13).value = str(duet_setting)
                                nickname=author.get("nickname")
                                if(nickname!=None):
                                    table.cell(nrows + index, 14).value = str(nickname)
                                favoriting_count=author.get("favoriting_count")
                                if(favoriting_count!=None):
                                    table.cell(nrows + index, 15).value = str(favoriting_count)
                                room_id=author.get("room_id")
                                if(room_id!=None):
                                    table.cell(nrows + index, 16).value = str(room_id)
                                followers_detail=author.get("followers_detail")
                                if(followers_detail!=None):
                                    table.cell(nrows + index, 17).value = str(followers_detail)
                                live_commerce=author.get("live_commerce")
                                if(live_commerce!=None):
                                    table.cell(nrows + index, 18).value = str(live_commerce)
                                is_star=author.get("is_star")
                                if(is_star!=None):
                                    table.cell(nrows + index, 19).value = str(is_star)
                                weibo_schema=author.get("weibo_schema")
                                if(weibo_schema!=None):
                                    table.cell(nrows + index, 20).value = str(weibo_schema)
                                school_poi_id=author.get("school_poi_id")
                                if(school_poi_id!=None):
                                    table.cell(nrows + index, 21).value = str(school_poi_id)
                                ins_id= author.get("ins_id")
                                if(ins_id!=None):
                                    table.cell(nrows + index, 22).value = str(ins_id)
                                signature=author.get("signature")
                                if(signature!=None):
                                    table.cell(nrows + index, 23).value = str(signature)
                                aweme_count=author.get("aweme_count")
                                if(aweme_count!=None):
                                    table.cell(nrows + index, 24).value = str(aweme_count)
                                live_verify=author.get("live_verify")
                                if(live_verify!=None):
                                    table.cell(nrows + index, 25).value = str(live_verify)
                                has_email=author.get("has_email")
                                if(has_email!=None):
                                    table.cell(nrows + index, 26).value = str(has_email)
                                cv_level=author.get("cv_level")
                                if(cv_level!=None):
                                    table.cell(nrows + index, 27).value = str(cv_level)
                                need_recommend=author.get("need_recommend")
                                if(need_recommend!=None):
                                    table.cell(nrows + index, 28).value = str(need_recommend)
                                weibo_name=author.get("weibo_name")
                                if(weibo_name!=None):
                                    table.cell(nrows + index, 29).value = str(weibo_name)
                                google_account=author.get("google_account")
                                if(google_account!=None):
                                    table.cell(nrows + index, 30).value = str(google_account)
                                birthday=author.get("birthday")
                                if(birthday!=None):
                                    table.cell(nrows + index, 31).value = str(birthday)
                                follower_count=author.get("follower_count")
                                if(follower_count!=None):
                                    table.cell(nrows + index, 32).value = str(follower_count)

                                constellation=author.get("constellation")
                                if(constellation!=None):
                                    table.cell(nrows + index, 34).value = str(constellation)
                                account_region=author.get("account_region")
                                if(account_region!=None):
                                    table.cell(nrows + index, 35).value = str(account_region)
                                enterprise_verify_reason=author.get("enterprise_verify_reason")
                                if(enterprise_verify_reason!=None):
                                    table.cell(nrows + index, 36).value = str(enterprise_verify_reason)
                                following_count=author.get("following_count")
                                if(following_count!=None):
                                    table.cell(nrows + index, 37).value = str(following_count)
                                youtube_channel_id=author.get("youtube_channel_id")
                                if(youtube_channel_id!=None):
                                    table.cell(nrows + index, 38).value = str(youtube_channel_id)
                                total_favorited=author.get("total_favorited")
                                if(total_favorited!=None):
                                    table.cell(nrows + index, 39).value = str(total_favorited)
                                weibo_url=author.get("weibo_url")
                                if(weibo_url!=None):
                                    table.cell(nrows + index, 40).value = str(weibo_url)
                                story_count=author.get("story_count")
                                if(story_count!=None):
                                    table.cell(nrows + index, 41).value = str(story_count)
                                bind_phone=author.get("bind_phone")
                                if(bind_phone!=None):
                                    table.cell(nrows + index, 42).value = str(bind_phone)
                                school_name=author.get("school_name")
                                if(school_name!=None):
                                    table.cell(nrows + index, 43).value = str(school_name)
                                gender=author.get("gender")
                                if(gender!=None):
                                    table.cell(nrows + index, 44).value = str(gender)
                                twitter_name=author.get("twitter_name")
                                if(twitter_name!=None):
                                    table.cell(nrows + index, 45).value = str(twitter_name)
                                location=author.get("location")
                                if(location!=None):
                                    table.cell(nrows + index, 46).value = str(location)
                                region=author.get("region")
                                if(region!=None):
                                    table.cell(nrows + index, 47).value = str(region)
                                youtube_channel_title=author.get("youtube_channel_title")
                                if(youtube_channel_title!=None):
                                    table.cell(nrows + index, 48).value = str(youtube_channel_title)


                            cha_list=eve.get("cha_list")
                            if(cha_list!=None):
                                cha_name=cha_list[0].get("cha_name")
                                if(cha_name!=None):
                                    table.cell(nrows + index, 49).value = str(cha_name)
                                view_count=cha_list[0].get("view_count")
                                if(view_count!=None):
                                    table.cell(nrows + index, 50).value = str(view_count)

                            share_info = eve.get("share_info")
                            if(share_info!=None):
                                share_title=share_info.get("share_title")
                                if(share_title!=None):
                                    table.cell(nrows + index, 51).value = str(share_title)
                                share_link_desc=share_info.get("share_link_desc")
                                if(share_link_desc!=None):
                                    table.cell(nrows + index, 52).value = str(share_link_desc)
                                share_url=share_info.get("share_url")
                                if(share_url!=None):
                                    table.cell(nrows + index, 53).value = str(share_url)
                                share_desc_info=share_info.get("share_desc_info")
                                if(share_desc_info!=None):
                                    table.cell(nrows + index, 54).value = str(share_desc_info)

                            video = eve.get("video")
                            if (video != None):
                                download_suffix_logo_addr=video.get("download_suffix_logo_addr")
                                if(download_suffix_logo_addr!=None):
                                    url_list=download_suffix_logo_addr.get("url_list")
                                    if(url_list!=None):
                                        mvurl = url_list[0]
                                        table.cell(nrows + index, 55).value = str(mvurl)
                                        try:
                                            opener = urllib.request.build_opener()
                                            opener.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.90 Safari/537.36")]
                                            urllib.request.install_opener(opener)
                                            urllib.request.urlretrieve(mvurl,'D:\\scratch\\douyin\\video\\1\\%s.mp4' % str(nrows+index))
                                        except  urllib.error.HTTPError as e:
                                            print(e.reason)
                                            print(e.code)
                                            print(e.headers)


                            music=eve.get("music")
                            if(music!=None):
                                music_title = music.get("title")
                                table.cell(nrows + index, 56).value = str(music_title)
                                music_anthor = music.get("anthor")
                                table.cell(nrows + index, 57).value = str(music_anthor)
                                temp = music.get("cover_medium")
                                music_pic = temp.get("url_list")
                                music_picture = music_pic[0]
                                table.cell(nrows + index, 58).value = str(music_picture)
                                music_playurls = music.get("play_url")
                                music_playurl = music_playurls.get("uri")
                                table.cell(nrows + index, 59).value = str(music_playurl)

                            statistics = eve.get("statistics")
                            comment_count = statistics.get("comment_count")
                            digg_count = statistics.get("digg_count")
                            download_count = statistics.get("download_count")
                            forward_count = statistics.get("forward_count")
                            lose_comment_count = statistics.get("lose_comment_count")
                            whatsapp_share_count = statistics.get("whatsapp_share_count")
                            play_count = statistics.get("play_count")
                            share_count = statistics.get("share_count")
                            lose_count = statistics.get("lose_count")
                            table.cell(nrows + index, 60).value = str(comment_count)
                            table.cell(nrows + index, 61).value = str(digg_count)
                            table.cell(nrows + index, 62).value = str(download_count)
                            table.cell(nrows + index, 63).value = str(forward_count)
                            table.cell(nrows + index, 64).value = str(lose_comment_count)
                            table.cell(nrows + index, 65).value = str(whatsapp_share_count)
                            table.cell(nrows + index, 66).value = str(play_count)
                            table.cell(nrows + index, 67).value = str(share_count)
                            table.cell(nrows + index, 68).value = str(lose_count)

                            time.sleep(0.5)  # 每下载一个视频后挂起十秒钟










                        index = index + 1
            data.save('douyin1.xlsx')

if __name__ == '__main__':
 douyin()