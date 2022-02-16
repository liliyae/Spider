import sys
import requests
import json
import urllib as UrlUtils
from bs4 import BeautifulSoup
import random
import ast
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import openpyxl

data = openpyxl.load_workbook('meituan.xlsx')
sheetnames = data.get_sheet_names()
table = data.active
nrows = table.max_row  # 获得行数

soup=BeautifulSoup(open('D:\\touristapp\\8.txt',encoding='utf-8'),features='html.parser')
print(soup.text)
# items 是一个 <listiterator object at 0x10a4b9950> 对象，不是一个list，但是可以循环遍历所有子节点。
items = soup.find(attrs={'class': 'common-list-main'}).children
projectList = []
i = nrows+1
for item in items:
    if item == '\n': continue
    # 获取需要的数据
    mechanurl = item.find_all('a')[1].text
    table.cell(i, 1).value = mechanurl

    j=2
    while(j<9):
        table.cell(i,j).value = item.find_all('span')[j-1].text
        print(table.cell(i,j).value)
        j=j+1
    #site=item.find(attrs={'class': 'item-site-info clearfix'}).string.strip()
    #price=item.find(attrs={'class': 'item-bottom-info clearfix'}).string.strip()
    print(mechanurl)
    print("aaaaaaaa")

    i = i+1
data.save("meituan.xlsx")

'''
    title = item.find(attrs={'class': 'title'}).string.strip()
    projectId = item.find(attrs={'class': 'subtitle'}).string.strip()
    projectType = item.find(attrs={'class': 'invest-item-subtitle'}).span.string
    percent = item.find(attrs={'class': 'percent'})
    state = 'Open'
    if percent is None: # 融资已完成
        percent = '100%'
        state = 'Finished'
        totalAmount = item.find(attrs={'class': 'project-info'}).span.string.strip()
        investedAmount = totalAmount
    else:
        percent = percent.string.strip()
        state = 'Open'
        decimalList = item.find(attrs={'class': 'decimal-wrap'}).find_all(attrs={'class': 'decimal'})
        totalAmount =  decimalList[0].string
        investedAmount = decimalList[1].string
    investState = item.find(attrs={'class': 'invest-item-type'})
    if investState != None:
        state = investState.string
    profitSpan = item.find(attrs={'class': 'invest-item-rate'}).find(attrs={'class': 'invest-item-profit'})
    profit1 = profitSpan.next.strip()
    profit2 = profitSpan.em.string.strip()
    profit = profit1 + profit2
    term = item.find(attrs={'class': 'invest-item-maturity'}).find(attrs={'class': 'invest-item-profit'}).string.strip()
    project = {
        'title': title,
        'projectId': projectId,
        'type': projectType,
        'percent': percent,
        'totalAmount': totalAmount,
        'investedAmount': investedAmount,
        'profit': profit,
        'term': term,
        'state': state
    }
    projectList.append(project)
'''