# -*- coding: UTF-8 -*-
# Only for research and not for public release #
import urllib.request
import os
import time
import re
import logging
import requests
from bs4 import BeautifulSoup
import urllib3
# json解析库,对应到lxml
import json
# 全局取消证书验证
import io
import sys
# 改变标准输出的默认编码
# utf-8中文乱码 有些字符写不进去
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import openpyxl


def kuishou_video():

    file_path = 'url.txt'  # 每个json的url储存文档
    if not os.path.isfile(file_path):
        raise TypeError(file_path + "dose not exit!")
    file_object = open(file_path).read().split('\n')


    #将下面这个url_first改为粘贴的POST后面的蓝色链接
    url_first="https://api3.gifshow.com/rest/nebula/tag/text/feed/hot?apptype=2&kcv=188&kpf=IPAD&net=_5&appver=2.3.2.78&kpn=NEBULA&mod=iPad11%2C1&c=a&sys=ios13.3&sh=2048&ver=2.3&isp=&did=758308F9-5C2A-4C4B-9F6A-0C3029769569&ud=1915020708&did_gt=1589275706348&sw=1536&browseType=3&egid=DFPAFC9D5D5D453AE928E32FBB0CB6C725C1C726C6F49ACD93CC7E865EC75790"

    page = 0  # 记录是第几页url
    video_count = 0  # 记录视频个数

    for video_url in file_object:
        data = openpyxl.load_workbook('excel_test.xlsx')
        print(data.get_named_ranges())  # 输出工作页索引范围
        print(data.get_sheet_names())  # 输出所有工作页的名称
        # 取第一张表
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        table = data.active
        print(table.title)  # 输出表名
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得行数
        index = 1

        page = page + 1
        print('page:', page)
        video_url = url_first+'&' + video_url
        print(video_url)
        req = urllib.request.urlopen(video_url).read()
        unicodestr = json.loads(req)
        user_list = unicodestr["feeds"]
        if user_list != None:
            for i in user_list:
                    mv = i.get("main_mv_urls_h265")
                    if (mv == None):
                        mv = i.get("main_mv_urls")
                    if (mv != None):
                        mvurl_list = mv[0]
                        mvurl = mvurl_list["url"]
                        urllib.request.urlretrieve(mvurl, 'D:\\video\\%s.mp4' % str(nrows+index))
                        table.cell(nrows + index, 1).value = str(mechant_url)
                        position = i.get("poi")
                        if (position != None):
                            city = position.get("city")
                            po_title = position["title"]
                            po_address = position["address"]
                            po_id = position["id"]
                            if (city != None):
                                table.cell(nrows + index, 2).value = str(city)
                            if (po_title != None):
                                table.cell(nrows + index, 3).value = str(po_title)
                            if (po_address != None):
                                table.cell(nrows + index, 4).value = str(po_address)
                            if (po_id != None):
                                table.cell(nrows + index, 5).value = str(po_id)
                        publishtime = i.get("time")
                        like_count = i.get("like_count")
                        share_count = i.get("share_count")
                        view_count = i.get("view_count")
                        comment_count = i.get("comment_count")
                        if(publishtime!=None):
                            table.cell(nrows + index, 6).value = str(publishtime)
                        if (like_count != None):
                            table.cell(nrows + index, 7).value = str(like_count)
                        if (share_count != None):
                            table.cell(nrows + index, 8).value =  str(share_count)
                        if (view_count != None):
                             table.cell(nrows + index, 9).value = str(view_count)
                        if (comment_count != None):
                            table.cell(nrows + index, 10).value = str(comment_count)
                        user_kuaiID = i.get("kwaiId")
                        user_sex = i.get("user_sex")
                        user_display_photo = i.get("headurls")
                        user_display_photo_url = user_display_photo[0].get("url")
                        user_name = i.get("user_name")
                        caption = i.get("caption")
                        if (user_kuaiID != None):
                            table.cell(nrows + index, 11).value = str(user_kuaiID)
                        if (user_name != None):
                            table.cell(nrows + index, 12).value = str(user_name)
                        if (user_sex != None):
                            table.cell(nrows + index, 13).value = str(user_sex)
                        if (user_display_photo_url != None):
                            table.cell(nrows + index, 14).value = str(user_display_photo_url)

                        music = i.get("music")
                        if (music != None):
                            music_user_information = music.get("user")
                            if (music_user_information != None):
                                music_user_kuaiID = music_user_information.get("kwaiId")
                                music_user_ID = music_user_information.get("user_id")
                                music_user_sex = music_user_information.get("user_sex")
                                music_user_display_photo = music_user_information.get("headurl")
                                music_user_name = music_user_information.get("user_name")
                                if (music_user_kuaiID != None):
                                    table.cell(nrows + index, 20).value = str(music_user_kuaiID)
                                if (music_user_sex != None):
                                    table.cell(nrows + index, 21).value = str(music_user_sex)
                                if (music_user_ID != None):
                                    table.cell(nrows + index, 22).value = str(music_user_ID)
                                if (music_user_display_photo != None):
                                    table.cell(nrows + index, 23).value = str(music_user_display_photo)
                                if (music_user_name != None):
                                    table.cell(nrows + index, 24).value = str(music_user_name)

                            bgm_url = music.get("audioUrls")
                            bgm_url = bgm_url[0].get("url")
                            bgm_name = music.get("name")
                            if (bgm_url != None):
                                table.cell(nrows + index, 25).value = str(bgm_url)
                            if (bgm_name != None):
                                table.cell(nrows + index, 26).value = str(bgm_name)

                        logger.info('page ' + str(page) + '\n' + 'index: ' + str(index) + '\n' + 'caption: ' + str(caption) + '\n')  # 打印信息到日志
                        index = index + 1
                        print("#################################################")
                        time.sleep(1)  # 每下载一个视频后挂起十秒钟

        data.save('excel_test.xlsx')

kuishou_video()



